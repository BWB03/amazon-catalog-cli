"""Tests for Seller Central listing lookup and diff helpers."""

import json

import openpyxl

from catalog.core.models import SellerListingDiffRequest, SellerListingFetchRequest
from catalog.core.seller_central import (
    HttpResult,
    build_reconciled_details_url,
    diff_seller_listing,
    fetch_seller_listing,
    resolve_cookie,
)


def _sample_response() -> dict:
    return {
        "detailPageListingResponse": {
            "brand#1.value": {"displayLabel": "Brand Name", "value": "Amazon Brand"},
            "item_name#1.value": {"displayLabel": "Title", "value": "Live Title"},
            "detail_page_primary_image_url": {"displayLabel": "", "value": "https://example.com/image.jpg"},
        },
        "detailPageListingResponseImsv3": json.dumps(
            {
                "brand": [{"value": "Amazon Brand"}],
                "item_name": [{"value": "Live Title"}],
            }
        ),
        "reconciledDataSummary": None,
        "debug": None,
    }


def _http_result(payload: dict, status_code: int = 200, content_type: str = "application/json") -> HttpResult:
    return HttpResult(
        status_code=status_code,
        headers={"Content-Type": content_type},
        body=json.dumps(payload).encode("utf-8"),
        url="https://sellercentral.amazon.com/abis/ajax/reconciledDetailsV2?asin=B000TEST01",
    )


def _write_clr(path):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Template"
    headers = ["Status", "Title", "SKU", "Product Type", "Product Id Type", "Product Id", "Brand Name"]
    for col, value in enumerate(headers, start=1):
        sheet.cell(row=4, column=col, value=value)
    values = ["Active", "Live Title", "SKU-1", "TEST_PRODUCT", "ASIN", "B000TEST01", "CLR Brand"]
    for col, value in enumerate(values, start=1):
        sheet.cell(row=7, column=col, value=value)
    workbook.save(path)


def test_build_reconciled_details_url_encodes_asin():
    assert build_reconciled_details_url("B000TEST01").endswith("asin=B000TEST01")


def test_resolve_cookie_precedence(tmp_path, monkeypatch):
    cookie_file = tmp_path / "cookie.txt"
    cookie_file.write_text("file-cookie", encoding="utf-8")
    monkeypatch.setenv("CATALOG_SELLER_CENTRAL_COOKIE", "env-cookie")

    assert resolve_cookie("explicit-cookie", str(cookie_file)) == "explicit-cookie"
    assert resolve_cookie(None, str(cookie_file)) == "file-cookie"
    assert resolve_cookie(None, None) == "env-cookie"


def test_fetch_requires_cookie(monkeypatch):
    monkeypatch.delenv("CATALOG_SELLER_CENTRAL_COOKIE", raising=False)

    response = fetch_seller_listing(SellerListingFetchRequest(asin="B000TEST01"))

    assert response.status == "auth_required"
    assert "cookie" in response.error.lower()


def test_fetch_parses_display_fields_and_imsv3(monkeypatch):
    def fake_http_get(url, cookie, timeout):
        return _http_result(_sample_response())

    monkeypatch.setattr("catalog.core.seller_central._http_get", fake_http_get)

    response = fetch_seller_listing(
        SellerListingFetchRequest(asin="B000TEST01", cookie="session-cookie")
    )

    assert response.status == "success"
    assert response.display_fields["brand#1.value"]["displayLabel"] == "Brand Name"
    assert response.parsed_imsv3["brand"][0]["value"] == "Amazon Brand"
    assert response.raw_response["reconciledDataSummary"] is None


def test_fetch_treats_html_as_auth_required(monkeypatch):
    def fake_http_get(url, cookie, timeout):
        return HttpResult(
            status_code=200,
            headers={"Content-Type": "text/html"},
            body=b"<html><title>Amazon Sign In</title></html>",
            url=url,
        )

    monkeypatch.setattr("catalog.core.seller_central._http_get", fake_http_get)

    response = fetch_seller_listing(
        SellerListingFetchRequest(asin="B000TEST01", cookie="expired-cookie")
    )

    assert response.status == "auth_required"
    assert "login" in response.error.lower()


def test_diff_compares_live_fields_to_matched_clr(tmp_path, monkeypatch):
    clr_path = tmp_path / "catalog.xlsx"
    _write_clr(clr_path)

    def fake_http_get(url, cookie, timeout):
        return _http_result(_sample_response())

    monkeypatch.setattr("catalog.core.seller_central._http_get", fake_http_get)

    response = diff_seller_listing(
        SellerListingDiffRequest(
            asin="B000TEST01",
            file=str(clr_path),
            cookie="session-cookie",
        )
    )

    assert response.status == "success"
    assert response.clr_match["sku"] == "SKU-1"
    assert response.value_mismatches == [
        {
            "field": "Brand Name",
            "amazon_field": "Brand Name",
            "clr_value": "CLR Brand",
            "amazon_value": "Amazon Brand",
        }
    ]
    assert "detail_page_primary_image_url" in response.amazon_only
    assert "Product Id" in response.clr_only
